# app.py — Image Extractor (Streamlit)
# Upload Excel/CSV with a 'thumbnail' URL column.
# Outputs a new Excel that embeds small image previews and adds 'thumbnail_dataurl' (base64) per row.

import io, base64, requests, urllib3
import pandas as pd
import streamlit as st
from PIL import Image
from requests.exceptions import SSLError
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------- Config ----------
TIMEOUT = 25
MAX_BYTES = 12_000_000
THUMB_PX = 96
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

st.set_page_config(page_title="Image Extractor", page_icon="🖼️", layout="centered")
st.title("🖼️ Image Extractor")
st.caption("Turns a products file with a 'thumbnail' URL column into an Excel with embedded previews and a 'thumbnail_dataurl' base64 column.")

# ---------- Inputs ----------
up = st.file_uploader("Upload products file (.xlsx/.xls/.csv)", type=["xlsx","xls","csv"])
limit = st.number_input("Max rows to process (0 = all)", min_value=0, value=0, step=100)
start_btn = st.button("Process")

# ---------- HTTP ----------
@st.cache_resource
def http_session():
    s = requests.Session()
    s.headers.update({
        "User-Agent": "Mozilla/5.0",
        "Accept": "image/avif,image/webp,image/apng,image/*,*/*;q=0.8",
        "Accept-Language": "en",
        "Cache-Control": "no-cache",
        "Referer": "",
    })
    return s

def fetch_image(url: str) -> bytes:
    if not isinstance(url, str) or not url.strip():
        return b""
    s = http_session()
    headers = {"Referer": url}
    try:
        r = s.get(url, timeout=TIMEOUT, stream=True, allow_redirects=True, headers=headers)
    except SSLError:
        r = s.get(url, timeout=TIMEOUT, stream=True, allow_redirects=True, headers=headers, verify=False)
    r.raise_for_status()
    data = r.content if r.content else r.raw.read(MAX_BYTES + 1)
    if not data and url.startswith("https://"):
        alt = "http://" + url[len("https://"):]
        try:
            r2 = s.get(alt, timeout=TIMEOUT, stream=True, allow_redirects=True, headers=headers)
            r2.raise_for_status()
            data = r2.content if r2.content else r2.raw.read(MAX_BYTES + 1)
        except Exception:
            return b""
    return data if data and len(data) <= MAX_BYTES else b""

# ---------- Imaging ----------
def to_png_thumb(img_bytes: bytes, max_px: int = THUMB_PX) -> bytes:
    im = Image.open(io.BytesIO(img_bytes)).convert("RGB")
    im.thumbnail((max_px, max_px))
    buf = io.BytesIO(); im.save(buf, format="PNG", optimize=True)
    return buf.getvalue()

def sniff_mime(img_bytes: bytes) -> str:
    try:
        fmt = Image.open(io.BytesIO(img_bytes)).format or "PNG"
    except Exception:
        fmt = "PNG"
    fmt = fmt.upper()
    if fmt in ("JPG","JPEG"): return "image/jpeg"
    if fmt == "PNG": return "image/png"
    if fmt == "GIF": return "image/gif"
    return "image/png"

def to_data_url(img_bytes: bytes) -> str:
    mime = sniff_mime(img_bytes)
    b64 = base64.b64encode(img_bytes).decode("ascii")
    return f"data:{mime};base64,{b64}"

# ---------- IO ----------
def read_any_table(f):
    fn = f.name.lower()
    if fn.endswith((".xlsx",".xls")): return pd.read_excel(f, engine="openpyxl")
    if fn.endswith(".csv"):           return pd.read_csv(f)
    raise ValueError("Use .xlsx/.xls/.csv")

def build_excel_with_images(df: pd.DataFrame, max_rows: int = 0) -> tuple[bytes, pd.DataFrame]:
    if "thumbnail" not in df.columns:
        raise ValueError("Missing required column: 'thumbnail'")

    df_out = df.copy()
    if max_rows and max_rows > 0:
        df_out = df_out.head(max_rows)

    if "thumbnail_dataurl" not in df_out.columns:
        df_out["thumbnail_dataurl"] = ""

    wb = Workbook(); ws = wb.active; ws.title = "Products"
    for r in dataframe_to_rows(df_out, index=False, header=True):
        ws.append(r)

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col_thumb = headers.index("thumbnail") + 1
    col_dataurl = headers.index("thumbnail_dataurl") + 1

    ws.insert_cols(col_dataurl + 1, amount=1)
    ws.cell(row=1, column=col_dataurl + 1, value="thumbnail_image_embedded")
    col_img = col_dataurl + 1
    ws.column_dimensions[ws.cell(row=1, column=col_img).column_letter].width = 18

    rows = ws.max_row - 1
    prog = st.progress(0.0, text="Downloading images…")

    logs = []
    ok = fail = 0
    for row_idx in range(2, ws.max_row + 1):
        url = str(ws.cell(row=row_idx, column=col_thumb).value or "").strip()
        dcell = ws.cell(row=row_idx, column=col_dataurl)
        try:
            img_bytes = fetch_image(url) if url else b""
            if img_bytes:
                dcell.value = to_data_url(img_bytes)
                thumb = to_png_thumb(img_bytes, THUMB_PX)
                xl_img = XLImage(io.BytesIO(thumb))
                ws.add_image(xl_img, ws.cell(row=row_idx, column=col_img).coordinate)
                ws.row_dimensions[row_idx].height = 80
                ok += 1
                logs.append({"row": row_idx-1, "status": "ok"})
            else:
                fail += 1
                logs.append({"row": row_idx-1, "status": "no-bytes"})
        except Exception as e:
            fail += 1
            logs.append({"row": row_idx-1, "status": f"error:{type(e).__name__}"})
        if rows > 0:
            prog.progress((row_idx-1)/rows, text=f"Processed {row_idx-1}/{rows}")

    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out.getvalue(), pd.DataFrame(logs)

# ---------- Action ----------
if start_btn:
    if not up:
        st.error("Upload a file first.")
    else:
        try:
            df = read_any_table(up)
            st.write({"rows": len(df), "columns": list(df.columns)})

            try:
                sample = df["thumbnail"].astype(str).head(6).tolist()
                st.caption("Preview of first URLs (browser view):")
                cols = st.columns(3)
                for i, u in enumerate(sample):
                    with cols[i % 3]:
                        st.image(u, caption=f"Row {i+1}", width="stretch")
            except Exception:
                pass

            buf, logs = build_excel_with_images(df, max_rows=int(limit))
            st.success("Done. Download and use in your main dashboard.")

            st.download_button(
                "Download preprocessed Excel",
                data=buf,
                file_name="products_with_images.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            if not logs.empty:
                embedded_ok = int((logs["status"] == "ok").sum())
                embedded_fail = len(logs) - embedded_ok
                st.write({"embedded_ok": embedded_ok, "embedded_fail": embedded_fail})
                st.dataframe(logs.head(50), width="stretch")
            else:
                st.write({"embedded_ok": 0, "embedded_fail": 0})

            st.info("Open the XLSX in Microsoft Excel desktop to see embedded images. Google Sheets/Numbers may hide them.")
        except Exception as e:
            st.error(f"Failed: {type(e).__name__}: {e}")
